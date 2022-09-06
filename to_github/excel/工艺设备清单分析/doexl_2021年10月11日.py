from PyQt5 import QtCore, QtWidgets
import sys


from PyQt5.QtWidgets import *
import re
from openpyxl.styles import PatternFill
from copy import copy
import openpyxl
import os

import winreg
# 获取桌面路径

def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')
    return winreg.QueryValueEx(key, 'Desktop')[0]


class Ui_MainWindow(QtWidgets.QMainWindow):

    def __init__(self):
        super(Ui_MainWindow, self).__init__()        
        self.setupUi(self)        
        # self.retranslateUi(self)
        self.fileName1 = ""
        self.setAcceptDrops(True) #文件拖入功能启用
        # self.setDragEnabled(True)
        #self.setupUi(MainWindow)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(413, 270)
        self.lineEdit = QtWidgets.QLineEdit(MainWindow)
        self.lineEdit.setGeometry(QtCore.QRect(30, 30, 291, 20))
        self.lineEdit.setObjectName("lineEdit")
        self.checkBox = QtWidgets.QCheckBox(MainWindow)
        self.checkBox.setGeometry(QtCore.QRect(-140, 230, 71, 16))
        self.checkBox.setObjectName("checkBox")
        self.open = QtWidgets.QPushButton(MainWindow)
        self.open.setGeometry(QtCore.QRect(330, 30, 75, 23))
        self.open.setObjectName("open")
        self.open.clicked.connect(self.openfile)
        self.open_2 = QtWidgets.QPushButton(MainWindow)
        self.open_2.setGeometry(QtCore.QRect(150, 200, 75, 23))
        self.open_2.setObjectName("open_2")
        self.open_2.clicked.connect(self.excel_ds)
        self.textBrowser = QtWidgets.QTextBrowser(MainWindow)
        self.textBrowser.setEnabled(False)
        self.textBrowser.setGeometry(QtCore.QRect(40, 80, 351, 91))
        self.textBrowser.setAcceptDrops(False)
        self.textBrowser.setToolTip("")
        self.textBrowser.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.textBrowser.setFrameShadow(QtWidgets.QFrame.Raised)
        self.textBrowser.setLineWidth(0)
        self.textBrowser.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self.textBrowser.setOpenLinks(False)
        self.textBrowser.setObjectName("textBrowser")

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        MainWindow.setTabOrder(self.checkBox, self.lineEdit)
        # MainWindow.setTabOrder(self.checkBox, self.lineEdit)
        # self.open.clicked.connect(self.openfile)
        
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "请将要处理的表格拖入此窗口，然后点击确定"))
        self.checkBox.setText(_translate("MainWindow", "CheckBox"))
        self.open.setText(_translate("MainWindow", "打开"))
        self.open_2.setText(_translate("MainWindow", "确定"))
        self.textBrowser.setHtml(_translate("MainWindow",
                                            "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                            "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                            "p, li { white-space: pre-wrap; }\n"
                                            "</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
                                            "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">说明：1、请确保标题行中有“规格”、“数量”关键字。</p>\n"
                                            "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">      2、请确设备清单位于表格第一个sheet。</p>\n"
                                            "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">      3、只能处理扩展名为xlsx的文档。</p>\n"
                                            "<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p>\n"
                                            "<p style=\"-qt-paragraph-type:empty; margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\"><br /></p></body></html>"))

    def openfile(self):
        # openfile_name = QFileDialog.getOpenFileName(self,'选择文件','','Excel files(*.xlsx , *.xls)')
        self.fileName1, filetype = QFileDialog.getOpenFileName(self, "选取文件", get_desktop(),"Excel Files (*.xlsx")  # 设置文件扩展名过滤,注意用双分号间隔
        print("打开文件名：", self.fileName1)
        self.lineEdit.setText(self.fileName1)

    def dragEnterEvent(self, event):
        # print("ddddddd")
        #self.fileName1 = os.path.dirname((event.mimeData().urls())[0].toLocalFile())     
        n =  event.mimeData().text()[8:]     
        if n.endswith(".xlsx"):
            self.fileName1 = n   
            print("打开文件名：", self.fileName1)
            self.lineEdit.setText(self.fileName1) 
            self.lineEdit.repaint()
        event.accept()
   

    def excel_ds(self):
        tablename = self.fileName1
        if tablename.endswith(".xlsx"):
            newtable = tablename.replace(".xlsx", "(电气).xlsx")
            print(newtable)
            s1 = newtable[:-5]
            print(s1)

            s = 1
            while os.path.exists(newtable):  

                newtable = s1 +"("+ str(s)+")"+".xlsx"

                print(s,newtable)
                s = s +1
                if s > 50:
                    break         
        
        book1 = openpyxl.load_workbook(tablename, data_only=True)
        # print(book1.sheetnames)
        w1 = book1[book1.sheetnames[0]]
        max_r = w1.max_row
        max_c = w1.max_column
        # print(max_r,max_c)
        gjc = ["数量", "序号", "名称", "规格", "材质", "备注", "品牌", "单位"]
        tj = [0]

        print("总行数：", max_r,"总列数：", max_c)

        for i in range(max_r):
            i = i +1
            tj.append(0)
            h = []
            for cell in w1[i]:
                # print(cell.value)
                h.append(cell.value)
            for j in h:
                for k in gjc:
                    match = re.search(k, str(j))
                    if match != None:
                        tj[i] = tj[i] + 1
        # print(tj)
        bth = tj.index(max(tj))
        print("标题行：", bth)

        # print(tj[bth])
        if tj[bth] > 3:
            y = []
            for cell in w1[bth]:
                # print(cell.value)
                y.append(cell.value)
            # y = sheet.row_values(bth)
            print(y)
            f_guige = 1000
            f_num = 1000
            for i in range(len(y)):
                if "规格" in str(y[i]):
                    print(y[i], "第", i, "列")
                    f_guige = i +1
                if "数量" == str(y[i]):
                    print(y[i], "第", i, "列")
                    f_num = i +1
            # ff = w1.cell(row = bth,column=f_guige).font
            # print(type(ff))
            # print(Font(name=u'宋体', size=12, bold=True, color='FF0000'))

            f_kw = max_c + 1
            w1.cell(row=bth, column=f_kw, value="设备功率").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_kw).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_kw).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_skw = max_c + 2
            w1.cell(row=bth, column=f_skw, value="总功率").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_skw).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_skw).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_bp = max_c + 3
            w1.cell(row=bth, column=f_bp, value="变频数量").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_bp).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_bp).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_zq = max_c + 4
            w1.cell(row=bth, column=f_zq, value="直启数量").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_zq).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_zq).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_11 = max_c + 5
            w1.cell(row=bth, column=f_11, value="<11kW").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_11).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_11).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_22 = max_c + 6
            w1.cell(row=bth, column=f_22, value="11-22kW").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_22).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_22).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_75 = max_c + 7
            w1.cell(row=bth, column=f_75, value="22-75kW").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_75).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_75).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            f_other = max_c + 8
            w1.cell(row=bth, column=f_other, value=">75kW").fill = PatternFill(fill_type='solid', start_color='00DB00')
            w1.cell(row=bth, column=f_other).font =copy(w1.cell(row = bth,column=f_guige).font)
            w1.cell(row=bth, column=f_other).alignment =copy(w1.cell(row = bth,column=f_guige).alignment)
            sum_kw = 0.0
            sum_zq = 0
            sum_bp = 0
            sum_11 = 0
            sum_22 = 0
            sum_75 = 0
            sum_other = 0
            if f_guige == 1000 or f_num == 1000:
                print("在标题行中未找到规格或数量列")
            else:
                for i in range(max_r):  # 获取excel中有多少行
                    i = i +1
                    if i > bth:
                        r1 = w1.cell(row=i,column=f_guige).value
                        # r1 = sheet.cell(i, f_guige).value
                        # print(i, r1)
                        t = w1.cell(row=i,column=f_num).value
                        # t = sheet.cell(i, f_num).value
                        r2 = str(r1)
                        r = r2.replace(" ", "")
                        r = r.replace(" ", "")
                        #print(r)
                        #print(np.fromstring(r1, dtype=np.uint8))
                        s1 = re.search("\d*\.?\d*[kK][wW]", r)
                        # print(i, r, s1)
                        if s1:
                            s2 = s1.group()
                            # print("S2:", s2)
                            s3 = float(s2[:len(s2) - 2])
                            w1.cell(row=i,column=f_kw,value=s3).fill = PatternFill(fill_type='solid', start_color='00DB00')
                            # print(t,type(t),type(s3))
                            w1.cell(row=i,column=f_skw,value=t*s3)
                            # ws.write(i, f_kw, s3)
                            # ws.write(i, f_skw, t * s3)
                            sum_kw = sum_kw + t * s3
                            k1 = re.search("变频", r)
                            if k1:
                                w1.cell(row=i,column=f_bp,value=t)
                                # ws.write(i, f_bp, t)
                                sum_bp = sum_bp + t
                            else:
                                w1.cell(row=i,column=f_zq,value=t)
                                # ws.write(i, f_zq, t)
                                sum_zq = sum_zq + t

                                if s3 <= 11:
                                    w1.cell(row=i,column=f_11,value=t)
                                    # ws.write(i, f_11, t)
                                    sum_11 = sum_11 + t
                                elif s3 <= 22 and s3 > 11:
                                    w1.cell(row=i,column=f_22,value=t)
                                    # ws.write(i, f_22, t)
                                    sum_22 = sum_22 + t
                                elif s3 <= 75 and s3 > 22:
                                    w1.cell(row=i,column=f_75,value=t)
                                    # ws.write(i, f_75, t)
                                    sum_75 = sum_75 + t
                                elif s3 > 75:
                                    w1.cell(row=i,column=f_other,value=t)
                                    # ws.write(i, f_other, t)
                                    sum_other = sum_other + t
                        else:
                            # w1.cell(row=i,column=f_kw).fill = PatternFill(fill_type='solid', start_color='00DB00')
                            #print("pass")
                            pass                
                w1.cell(row=max_r+1,column=f_skw-1,value="合计：")  
                w1.cell(row=max_r+1,column=f_skw,value=sum_kw)  
                w1.cell(row=max_r+1,column=f_bp,value=sum_bp) 
                w1.cell(row=max_r+1,column=f_zq,value=sum_zq) 
                w1.cell(row=max_r+1,column=f_11,value=sum_11) 
                w1.cell(row=max_r+1,column=f_22,value=sum_22) 
                w1.cell(row=max_r+1,column=f_75,value=sum_75) 
                w1.cell(row=max_r+1,column=f_other,value=sum_other) 

                # ws.write(i + 1, f_skw, sum_kw)
                # ws.write(i + 1, f_bp, sum_bp)
                # ws.write(i + 1, f_zq, sum_zq)
                # ws.write(i + 1, f_11, sum_11)
                # ws.write(i + 1, f_22, sum_22)
                # ws.write(i + 1, f_75, sum_75)
                # ws.write(i + 1, f_other, sum_other)
                # #print("测试点02")

        else:
            print("未找到标题行")
        try:
            book1.save(newtable)
            print("新文件保存成功")
        except:
            print("文件保存失败")   


if __name__ == "__main__":
    myapp = QApplication(sys.argv)
    myDlg = Ui_MainWindow()
    myDlg.show()
    sys.exit(myapp.exec_())


