from openpyxl import  Workbook 
from openpyxl import load_workbook
# 实例化
import re
tablename = "C:/Users/VM/Desktop/新乡浓缩液项目清单20211008（电气）.xlsx"


book1 = load_workbook(tablename)
w1 = book1[book1.sheetnames[0]]
max_r = w1.max_row
max_c = w1.max_column
print(max_r,max_c)
gjc = ["数量", "序号", "名称", "规格", "材质", "备注", "品牌", "单位"]
tj = [0]

print("总行数：", max_r)
for i in range(max_r):
    i = i +1
    tj.append(0)
    h = []
    for cell in w1[i]:
        # print(cell.value)
        h.append(cell.value)
    for j in h:
        for k in gjc:
            match = re.search(k, str(j))
            if match != None:
                tj[i] = tj[i] + 1
# print(tj)
bth = tj.index(max(tj))
print("标题行：", bth)

# print(tj[bth])
if tj[bth] > 3:
    y = []
    for cell in w1[bth]:
        # print(cell.value)
        y.append(cell.value)
    # y = sheet.row_values(bth)
    print(y)
    f_guige = 1000
    f_num = 1000
    for i in range(len(y)):
        if "规格" in str(y[i]):
            print(y[i], "第", i, "列")
            f_guige = i +1
        if "数量" == str(y[i]):
            print(y[i], "第", i, "列")
            f_num = i +1
    w1.cell(row=1, column=2, value=10)

    f_kw = max_c + 1
    w1.cell(row=bth, column=f_kw, value="设备功率")
    f_skw = max_c + 2
    w1.cell(row=bth, column=f_skw, value="总功率")
    f_bp = max_c + 3
    w1.cell(row=bth, column=f_bp, value="变频数量")
    f_zq = max_c + 4
    w1.cell(row=bth, column=f_zq, value="直启数量")
    f_11 = max_c + 5
    w1.cell(row=bth, column=f_11, value="<11kW")
    f_22 = max_c + 6
    w1.cell(row=bth, column=f_22, value="11-22kW")
    f_75 = max_c + 7
    w1.cell(row=bth, column=f_75, value="22-75kW")
    f_other = max_c + 8
    w1.cell(row=bth, column=f_other, value=">75kW")
    sum_kw = 0.0
    sum_zq = 0
    sum_bp = 0
    sum_11 = 0
    sum_22 = 0
    sum_75 = 0
    sum_other = 0
    if f_guige == 1000 or f_num == 1000:
        print("在标题行中未找到规格或数量列")
    else:
        for i in range(max_r):  # 获取excel中有多少行
            i = i +1
            if i > bth:
                r1 = w1.cell(row=i,column=f_guige).value
                # r1 = sheet.cell(i, f_guige).value
                # print(i, r1)
                t = w1.cell(row=i,column=f_num).value
                # t = sheet.cell(i, f_num).value
                r2 = str(r1)
                r = r2.replace(" ", "")
                r = r.replace(" ", "")
                #print(r)
                #print(np.fromstring(r1, dtype=np.uint8))
                s1 = re.search("\d*\.?\d*[kK][wW]", r)
                print(i, r, s1)
                if s1:
                    s2 = s1.group()
                    # print("S2:", s2)
                    s3 = float(s2[:len(s2) - 2])
                    w1.cell(row=i,column=f_kw,value=s3)
                    w1.cell(row=i,column=f_skw,value=t*s3)
                    # ws.write(i, f_kw, s3)
                    # ws.write(i, f_skw, t * s3)
                    sum_kw = sum_kw + t * s3
                    k1 = re.search("变频", r)
                    if k1:
                        w1.cell(row=i,column=f_bp,value=t)
                        # ws.write(i, f_bp, t)
                        sum_bp = sum_bp + t
                    else:
                        w1.cell(row=i,column=f_zq,value=t)
                        # ws.write(i, f_zq, t)
                        sum_zq = sum_zq + t

                        if s3 <= 11:
                            w1.cell(row=i,column=f_11,value=t)
                            # ws.write(i, f_11, t)
                            sum_11 = sum_11 + t
                        elif s3 <= 22 and s3 > 11:
                            w1.cell(row=i,column=f_22,value=t)
                            # ws.write(i, f_22, t)
                            sum_22 = sum_22 + t
                        elif s3 <= 75 and s3 > 22:
                            w1.cell(row=i,column=f_75,value=t)
                            # ws.write(i, f_75, t)
                            sum_75 = sum_75 + t
                        elif s3 > 75:
                            w1.cell(row=i,column=f_other,value=t)
                            # ws.write(i, f_other, t)
                            sum_other = sum_other + t
                else:
                    #print("pass")
                    pass
        w1.cell(row=max_r+1,column=f_skw,value=sum_kw)  
        w1.cell(row=max_r+1,column=f_bp,value=sum_bp) 
        w1.cell(row=max_r+1,column=f_zq,value=sum_zq) 
        w1.cell(row=max_r+1,column=f_11,value=sum_11) 
        w1.cell(row=max_r+1,column=f_22,value=sum_22) 
        w1.cell(row=max_r+1,column=f_75,value=sum_75) 
        w1.cell(row=max_r+1,column=f_other,value=sum_other) 

        # ws.write(i + 1, f_skw, sum_kw)
        # ws.write(i + 1, f_bp, sum_bp)
        # ws.write(i + 1, f_zq, sum_zq)
        # ws.write(i + 1, f_11, sum_11)
        # ws.write(i + 1, f_22, sum_22)
        # ws.write(i + 1, f_75, sum_75)
        # ws.write(i + 1, f_other, sum_other)
        # #print("测试点02")

else:
    print("未找到标题行")

book1.save("C:/Users/VM/Desktop/新乡浓缩液项目清单20211008（电气1）.xlsx")



# print(w1.max_row)
# print(w1.max_column)