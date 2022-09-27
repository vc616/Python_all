# 在NC导出的价格表格中，查找最新时间采购价格，并在该行最后一列标注1
# 需要设置表格名称

wlbm = 8     #物料编码所在的列
scsj = 36    #生成时间所在的列

name = "J:\\Users\\vc\\Desktop\\1200价格.xlsx"
newname= "J:\\Users\\vc\\Desktop\\1200价格1.xlsx"

import openpyxl
import datetime
# 加载 excel 文件
wb = openpyxl.load_workbook(name)

# 得到sheet对象
ws = wb['历史价格查询']
max_r = ws.max_row
max_c = ws.max_column
print(max_r)
# print(ws.cell(row = 9, column= 1).value)
# print(ws.cell(row = 10, column= 1).value)

b = []

for i in  range(max_r):
    if i >0 and i < max_r :
        q = ws.cell(row = i+1, column= wlbm).value
        if q != None:
            b.append(q)
print(b)
b1  = []
for i in b:
    if (i not in b1):
        b1.append(i)
print(b1)

for j in b1:
    print(j)
    t0 = datetime.datetime(2010, 4, 9, 0, 0)
    for i in  range(max_r):        
        if j == ws.cell(row = i+1, column= wlbm ).value:
            print(i)
            t=datetime.datetime.strptime(ws.cell(row = i+1, column= scsj).value,"%Y-%m-%d")
            if t > t0 :
                mx = i+1
                t0 =t
    print(t,mx)
    ws.cell(row = mx,column= max_c+1,value= 1)
wb.save(newname)


